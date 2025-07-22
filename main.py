
import os
from flask import Flask, request, jsonify
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import io

app = Flask(__name__)

def parse_excel(file_stream):
    try:
        workbook = load_workbook(filename=io.BytesIO(file_stream), data_only=True)
        if "Užsakymas" not in workbook.sheetnames:
            return {"error": "Sheet 'Užsakymas' not found in Excel file."}, None

        ws = workbook["Užsakymas"]

        customer = {
            "Klientas": ws["C4"].value or "",
            "El. paštas": ws["C5"].value or "",
            "Tel. nr.": ws["C6"].value or "",
            "Adresas": ws["C7"].value or "",
            "Miestas": ws["C8"].value or "",
            "Pastabos": ws["C9"].value or ""
        }

        order_info = {
            "Užsakymo data": ws["F4"].value.strftime("%Y-%m-%d") if isinstance(ws["F4"].value, datetime) else "",
            "Užsakymo nr.": str(ws["F5"].value or ""),
            "Gamybos terminas": ws["F6"].value.strftime("%Y-%m-%d") if isinstance(ws["F6"].value, datetime) else "",
            "Vadybininkas": ws["F7"].value or "",
            "Būsena": ws["F8"].value or "Naujas"
        }

        order_lines = []
        start_row = 14
        for row in ws.iter_rows(min_row=start_row, max_col=20):
            if all(cell.value is None for cell in row[:4]):
                continue  # skip empty rows
            line = {
                "Pozicija": row[0].value or "",
                "Patalpa": row[1].value or "",
                "Plotis": row[2].value or "",
                "Aukštis": row[3].value or "",
                "Kiekis": row[4].value or 1,
                "Valdymas": row[5].value or "",
                "Spalva": row[6].value or "",
                "Pakavimas": row[7].value or "",
                "Pastabos": row[8].value or ""
            }
            order_lines.append(line)

        return {
            "klientas": customer,
            "uzsakymas": order_info,
            "eilutes": order_lines
        }, None
    except Exception as e:
        return None, str(e)


@app.route("/")
def index():
    return "Excel to Airtable API veikia."

@app.route("/process", methods=["POST"])
def process_excel():
    if "file" not in request.files:
        return jsonify({"error": "Trūksta 'file' parametro"}), 400

    file = request.files["file"]
    content = file.read()

    data, error = parse_excel(content)
    if error:
        return jsonify({"error": error}), 500

    return jsonify(data)

if __name__ == "__main__":
    app.run(debug=True)
